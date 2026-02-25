"""
Экспорт сметы в Excel
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

_JUSTIFICATIONS_PATH = Path(__file__).parent.parent / "data" / "normative_justifications.json"

def _load_justifications() -> dict:
    """Загрузить обоснования объёмов из JSON."""
    try:
        with open(_JUSTIFICATIONS_PATH, encoding='utf-8') as f:
            return json.load(f).get("template_justifications", {})
    except Exception:
        return {}


def export_to_excel(estimate, filename: str = None) -> Path:
    """Экспортировать смету в Excel"""
    
    if filename is None:
        filename = f"Смета_{estimate.project_name}_{estimate.date_created}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета ИГИ"
    
    # Стили
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    money_format = '#,##0'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    subtotal_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    total_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    # Заголовок сметы
    ws.merge_cells('A1:G1')
    ws['A1'] = "ЛОКАЛЬНАЯ СМЕТА"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"на инженерно-геологические изыскания"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Информация о проекте
    row = 4
    project_info = [
        ("Проект:", estimate.project_name),
        ("Шифр:", estimate.project_code or "-"),
        ("Объект:", estimate.object_name or "-"),
        ("Заказчик:", estimate.customer or "-"),
        ("Подрядчик:", estimate.contractor or "-"),
        ("Дата:", estimate.date_created),
        ("Базовый город:", getattr(estimate, 'base_city', 'г. Санкт-Петербург')),
        ("Регион производства работ:", getattr(estimate, 'work_region', '-')),
        ("Расстояние до объекта:", f"{getattr(estimate, 'distance_km', '-')} км"),
        ("Индекс пересчёта:", f"{float(estimate.price_index):.2f}"),
    ]
    
    for label, value in project_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    
    # Заголовок таблицы
    headers = ["№ п/п", "Наименование работ", "Ед. изм.", "Кол-во", "Обоснование", "Расчёт", "Стоимость, руб."]
    col_widths = [8, 50, 10, 10, 20, 25, 18]
    
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = width
    
    row += 1
    start_data_row = row
    
    # Группировка по категориям
    categories = {
        "field": {"name": "ПОЛЕВЫЕ РАБОТЫ", "items": []},
        "laboratory": {"name": "ЛАБОРАТОРНЫЕ РАБОТЫ", "items": []},
        "office": {"name": "КАМЕРАЛЬНЫЕ РАБОТЫ", "items": []}
    }
    
    for item in estimate.items:
        # Определяем категорию по коду
        if item.code.startswith(("01", "02", "03", "04")):
            categories["field"]["items"].append(item)
        elif item.code.startswith(("05", "06", "07")):
            categories["laboratory"]["items"].append(item)
        else:
            categories["office"]["items"].append(item)
    
    item_num = 1
    
    for cat_key, cat_data in categories.items():
        if not cat_data["items"]:
            continue
        
        # Заголовок раздела
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = cat_data["name"]
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = header_fill
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
        
        # Позиции
        for item in cat_data["items"]:
            ws.cell(row=row, column=1, value=item_num).border = thin_border
            # Наименование
            ws.cell(row=row, column=2, value=item.name).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            
            # Ед. изм.
            ws.cell(row=row, column=3, value=item.unit).border = thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            # Кол-во
            qty_cell = ws.cell(row=row, column=4, value=float(item.quantity))
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(horizontal='right')
            
            # Обоснование (НЗ, таблица)
            ref_text = f"НЗ №281/пр, {item.table_ref}" if item.table_ref else "НЗ №281/пр"
            ref_cell = ws.cell(row=row, column=5, value=ref_text)
            ref_cell.border = thin_border
            ref_cell.alignment = Alignment(wrap_text=True, vertical='center')

            
            # Расчёт (Формула)
            formula_text = item.formula if item.formula else f"{float(item.base_cost):,.0f} x {float(item.quantity):,.1f}"
            calc_cell = ws.cell(row=row, column=6, value=formula_text)
            calc_cell.border = thin_border
            calc_cell.alignment = Alignment(horizontal='right')
            
            # Стоимость
            total_cell = ws.cell(row=row, column=7, value=float(item.total_cost))
            total_cell.border = thin_border
            total_cell.number_format = money_format
            
            item_num += 1
            row += 1
        
        # Подитог раздела
        subtotal = sum(float(item.total_cost) for item in cat_data["items"])
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Итого по разделу «{cat_data['name'].lower()}»:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).fill = subtotal_fill
        
        subtotal_cell = ws.cell(row=row, column=7, value=subtotal)
        subtotal_cell.font = Font(bold=True)
        subtotal_cell.number_format = money_format
        subtotal_cell.border = thin_border
        subtotal_cell.fill = subtotal_fill
        row += 1
    
    # -------------------------------------------------------------
    # ИТОГИ
    # -------------------------------------------------------------
    
    # 1. Базовые затраты (Сумма всех работ)
    base_total = sum(float(item.total_cost) for item in estimate.items)
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ИТОГО базовые затраты (СП + СЛ + СК):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = subtotal_fill
        
    total_cell = ws.cell(row=row, column=7, value=base_total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = money_format
    total_cell.border = thin_border
    total_cell.fill = subtotal_fill
    
    # 2. Дополнительные затраты (построчно)
    dz_sum = 0
    if estimate.additional_costs:
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "Дополнительные затраты:"
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        ws[f'A{row}'].border = thin_border
        row += 1
        
        dz_item_num = 1
        for cost in estimate.additional_costs:
            # 1. № п/п (можно продолжить или свой)
            ws.cell(row=row, column=1, value=f"ДЗ-{dz_item_num}").border = thin_border
            
            # 2. Наименование
            ws.cell(row=row, column=2, value=cost.get('name', 'ДЗ')).border = thin_border
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            
            # 3. Ед. изм.
            ws.cell(row=row, column=3, value="-").border = thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            # 4. Кол-во
            ws.cell(row=row, column=4, value="-").border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            
            # 5. Обоснование
            ws.cell(row=row, column=5, value=cost.get('basis', '-')).border = thin_border
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
            
            # 6. Расчёт
            ws.cell(row=row, column=6, value=cost.get('formula', '-')).border = thin_border
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            
            # 7. Стоимость
            val = round(float(cost.get('value', 0)), 2)
            dz_sum += val
            val_cell = ws.cell(row=row, column=7, value=val)
            val_cell.number_format = money_format
            val_cell.border = thin_border
            
            row += 1
            dz_item_num += 1
            
    # 3. Итого с учетом ДЗ
    total_with_dz = round(base_total + dz_sum, 2)
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ИТОГО с учетом дополнительных затрат:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = subtotal_fill

    total_dz_cell = ws.cell(row=row, column=7, value=total_with_dz)
    total_dz_cell.font = Font(bold=True)
    total_dz_cell.number_format = money_format
    total_dz_cell.border = thin_border
    total_dz_cell.fill = subtotal_fill
    
    # 4. С индексом пересчета
    idx = float(estimate.price_index)
    total_indexed = round(total_with_dz * idx, 2)
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"ИТОГО с индексом пересчёта ({idx:.2f}):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = total_fill
    
    total_indexed_cell = ws.cell(row=row, column=7, value=total_indexed)
    total_indexed_cell.font = Font(bold=True)
    total_indexed_cell.number_format = money_format
    total_indexed_cell.border = thin_border
    total_indexed_cell.fill = total_fill

    # 5. Коэффициент договорной цены
    k_contract = float(estimate.contract_coefficient)
    if k_contract != 1.0:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Коэффициент договорной цены ({k_contract:.3f}):"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border
        
        # Здесь выводим просто сумму
        # Но если по образцу, то это применяется к итогу
        final_total = round(total_indexed * k_contract, 2)
        
        k_cell = ws.cell(row=row, column=7, value=final_total)
        k_cell.font = Font(bold=True)
        k_cell.number_format = money_format
        k_cell.border = thin_border
    else:
        final_total = total_indexed

    # ВСЕГО ПО СМЕТЕ
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "ВСЕГО по смете:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = total_fill
    
    final_cell = ws.cell(row=row, column=7, value=final_total)
    final_cell.font = Font(bold=True, size=12)
    final_cell.number_format = money_format
    final_cell.border = thin_border
    final_cell.fill = total_fill
    
    # Подпись
    row += 3
    ws[f'A{row}'] = "Составил: __________________ / __________________ /"
    row += 2
    ws[f'A{row}'] = "Проверил: __________________ / __________________ /"

    # =========================================================
    # ВКЛАДКА 2: ОБОСНОВАНИЕ
    # =========================================================
    ws2 = wb.create_sheet("Обоснование")

    # Цвета
    hdr2_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sec_fill   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    dz_fill    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    even_fill  = PatternFill(start_color="F5F9FF", end_color="F5F9FF", fill_type="solid")
    total2_fill= PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    hdr2_font  = Font(bold=True, color="FFFFFF", size=10)
    sec_font   = Font(bold=True, size=10)
    normal_font= Font(size=10)
    ref_font   = Font(size=9, italic=True, color="1F4E79")

    wrap_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
    wrap_left   = Alignment(wrap_text=True, vertical='center', horizontal='left')
    wrap_right  = Alignment(wrap_text=True, vertical='center', horizontal='right')

    thin2 = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Загружаем обоснования объёмов
    all_justifications = _load_justifications()
    template_id = getattr(estimate, 'template_id', None)
    qty_justifications = all_justifications.get(template_id, {})

    # Ширина колонок вкладки 2 (11 колонок)
    col2_widths = {
        'A': 5,   # №
        'B': 44,  # Наименование
        'C': 7,   # Ед.
        'D': 6,   # Кол-во
        'E': 13,  # Норм. база (стоимость)
        'F': 11,  # ПЗ (цена)
        'G': 18,  # Коэффициенты
        'H': 22,  # Формула расчёта
        'I': 10,  # Стоимость
        'J': 22,  # Норм. основание объёма (СП, ГОСТ)
        'K': 42,  # Обоснование объёма
    }
    for col_letter, width in col2_widths.items():
        ws2.column_dimensions[col_letter].width = width

    # Заливка для колонок обоснования
    qty_fill   = PatternFill(start_color="EBF3E8", end_color="EBF3E8", fill_type="solid")
    qty_miss   = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")

    # ----- Заголовок листа -----
    ws2.merge_cells('A1:K1')
    ws2['A1'] = "ПОЯСНИТЕЛЬНАЯ ЗАПИСКА К СМЕТЕ (Обоснование стоимости и объёмов работ)"
    ws2['A1'].font = Font(bold=True, size=13)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 22

    ws2.merge_cells('A2:K2')
    ws2['A2'] = "Нормативная база: НЗ №281/пр (Приказ Минстроя РФ от 12.05.2025 № 281/пр) | СП 446.1325800.2019 | СП 341.1325800.2017 | СП 47.13330.2016 | ГОСТ 20522-2012"
    ws2['A2'].font = Font(italic=True, size=9, color="444444")
    ws2['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws2.row_dimensions[2].height = 24

    ws2.merge_cells('A3:K3')
    ws2['A3'] = (
        f"Проект: {estimate.project_name}   |   "
        f"Объект: {estimate.object_name or '—'}   |   "
        f"Регион: {getattr(estimate, 'work_region', '—') or '—'}   |   "
        f"Шаблон: {template_id or '—'}   |   "
        f"Дата: {estimate.date_created}"
    )
    ws2['A3'].font = Font(size=9, bold=True)
    ws2['A3'].alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[3].height = 16

    r2 = 5  # начальная строка данных

    # ----- Заголовок таблицы (11 колонок) -----
    headers2 = [
        "№", "Наименование работ", "Ед.", "Кол-во",
        "Норм. база\n(НЗ №281/пр)", "ПЗ, руб.", "Коэф-ты",
        "Формула расчёта", "Стоимость,\nруб.",
        "Норм. основание\nобъёма (СП/ГОСТ)",
        "Обоснование объёма работ"
    ]
    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(row=r2, column=ci, value=h)
        c.font = hdr2_font
        c.fill = hdr2_fill if ci <= 9 else PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        c.alignment = wrap_center
        c.border = thin2
    ws2.row_dimensions[r2].height = 32
    r2 += 1

    def w2_border_row(row_idx, ncols=11):
        for ci in range(1, ncols + 1):
            ws2.cell(row=row_idx, column=ci).border = thin2

    def w2_section(row_idx, title):
        ws2.merge_cells(f'A{row_idx}:K{row_idx}')
        c = ws2.cell(row=row_idx, column=1, value=title)
        c.font = sec_font
        c.fill = sec_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        w2_border_row(row_idx)
        ws2.row_dimensions[row_idx].height = 18

    # ----- Пояснение для каждого коэффициента -----
    def coef_note(item):
        notes = []
        k1 = float(item.k1) if hasattr(item, 'k1') else 1.0
        k2 = float(item.k2) if hasattr(item, 'k2') else 1.0
        k3 = float(item.k3) if hasattr(item, 'k3') else 1.0
        kc = float(item.climate_coef) if hasattr(item, 'climate_coef') else 1.0
        if k1 != 1.0: notes.append(f"К1={k1:.2f} (категория ИГУ)")
        if k2 != 1.0: notes.append(f"К2={k2:.2f} (климат. зона)")
        if k3 != 1.0: notes.append(f"К3={k3:.2f} (доп. коэф.)")
        if kc != 1.0: notes.append(f"Кклим={kc:.2f}")
        return "; ".join(notes) if notes else "—"

    # ----- Группировка -----
    cats2 = {
        "field":      {"name": "ПОЛЕВЫЕ РАБОТЫ",      "items": []},
        "laboratory": {"name": "ЛАБОРАТОРНЫЕ РАБОТЫ", "items": []},
        "office":     {"name": "КАМЕРАЛЬНЫЕ РАБОТЫ",  "items": []},
    }
    for item in estimate.items:
        if item.code.startswith(("01","02","03","04")):
            cats2["field"]["items"].append(item)
        elif item.code.startswith(("05","06","07")):
            cats2["laboratory"]["items"].append(item)
        else:
            cats2["office"]["items"].append(item)

    item_num2 = 1
    row_shade = False

    for cat_key, cat_data in cats2.items():
        if not cat_data["items"]:
            continue

        w2_section(r2, f"Раздел: {cat_data['name']}")
        r2 += 1

        for item in cat_data["items"]:
            ws2.row_dimensions[r2].height = 52
            fill_r = even_fill if row_shade else PatternFill()
            row_shade = not row_shade

            # № (A)
            c = ws2.cell(row=r2, column=1, value=item_num2)
            c.font = normal_font; c.alignment = wrap_center; c.border = thin2; c.fill = fill_r

            # Наименование (B)
            c = ws2.cell(row=r2, column=2, value=item.name)
            c.font = normal_font; c.alignment = wrap_left; c.border = thin2; c.fill = fill_r

            # Ед. (C)
            c = ws2.cell(row=r2, column=3, value=item.unit)
            c.font = normal_font; c.alignment = wrap_center; c.border = thin2; c.fill = fill_r

            # Кол-во (D)
            c = ws2.cell(row=r2, column=4, value=float(item.quantity))
            c.font = normal_font; c.alignment = wrap_right; c.border = thin2; c.fill = fill_r

            # Нормативная база (E)
            ref_text = f"НЗ №281/пр\n{item.table_ref}" if item.table_ref else "НЗ №281/пр"
            c = ws2.cell(row=r2, column=5, value=ref_text)
            c.font = ref_font; c.alignment = wrap_center; c.border = thin2; c.fill = fill_r

            # ПЗ базовая цена (F) — для рекогносцировки разбиваем
            pz_fixed = float(item.pz1p_fixed) if hasattr(item, 'pz1p_fixed') else 0
            if pz_fixed > 0:
                pz_text = f"ПЗ1п: {pz_fixed:,.0f}\nПЗ2п: {float(item.base_cost):,.0f}"
            else:
                pz_text = f"{float(item.base_cost):,.0f}"
            c = ws2.cell(row=r2, column=6, value=pz_text)
            c.font = normal_font; c.alignment = wrap_right; c.border = thin2; c.fill = fill_r

            # Коэффициенты / пояснение (G)
            cn = coef_note(item)
            c = ws2.cell(row=r2, column=7, value=cn)
            c.font = Font(size=9); c.alignment = wrap_left; c.border = thin2; c.fill = fill_r

            # Формула (H)
            formula_txt = item.formula if item.formula else (
                f"ПЗ1п({pz_fixed:,.0f}) + ПЗ2п({float(item.base_cost):,.0f}) × {float(item.quantity):.0f}"
                if pz_fixed > 0
                else f"{float(item.base_cost):,.0f} × {float(item.quantity):.1f}"
            )
            c = ws2.cell(row=r2, column=8, value=formula_txt)
            c.font = Font(size=9); c.alignment = wrap_right; c.border = thin2; c.fill = fill_r

            # Стоимость (I)
            c = ws2.cell(row=r2, column=9, value=float(item.total_cost))
            c.font = Font(bold=True, size=10)
            c.number_format = money_format
            c.alignment = wrap_right; c.border = thin2; c.fill = fill_r

            # Нормативное основание объёма (J) и Обоснование (K)
            work_id = getattr(item, 'work_id', item.code)
            jdata = qty_justifications.get(work_id, {})
            qty_ref  = jdata.get('qty_basis', '')
            qty_note = jdata.get('qty_note', '')
            has_just = bool(qty_ref or qty_note)
            j_fill = qty_fill if has_just else qty_miss

            cj = ws2.cell(row=r2, column=10, value=qty_ref or '—')
            cj.font = Font(size=8, italic=True, color="1A5C2A" if has_just else "AA0000")
            cj.alignment = wrap_left; cj.border = thin2; cj.fill = j_fill

            ck = ws2.cell(row=r2, column=11, value=qty_note or ('Нет данных' if not template_id else 'Нет обоснования для данного шаблона'))
            ck.font = Font(size=9, color="1A5C2A" if has_just else "AA0000")
            ck.alignment = wrap_left; ck.border = thin2; ck.fill = j_fill

            item_num2 += 1
            r2 += 1

        # Подитог раздела
        subtotal2 = sum(float(i.total_cost) for i in cat_data["items"])
        ws2.merge_cells(f'A{r2}:H{r2}')
        c = ws2.cell(row=r2, column=1, value=f"Итого по разделу «{cat_data['name'].lower()}»:")
        c.font = Font(bold=True, size=10); c.fill = subtotal_fill
        c.alignment = Alignment(horizontal='right', vertical='center')
        w2_border_row(r2)
        for ci2 in range(1, 12):
            ws2.cell(row=r2, column=ci2).fill = subtotal_fill
        c9 = ws2.cell(row=r2, column=9, value=subtotal2)
        c9.font = Font(bold=True, size=10); c9.number_format = money_format
        c9.alignment = wrap_right; c9.fill = subtotal_fill; c9.border = thin2
        ws2.row_dimensions[r2].height = 18
        r2 += 1

    r2 += 1  # отступ

    # =========================================================
    # ДЗ (Дополнительные затраты) — расширенное обоснование
    # =========================================================
    if estimate.additional_costs:
        ws2.merge_cells(f'A{r2}:K{r2}')
        c = ws2.cell(row=r2, column=1,
                     value="ДОПОЛНИТЕЛЬНЫЕ ЗАТРАТЫ (ДЗ) — обоснование")
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        w2_border_row(r2)
        ws2.row_dimensions[r2].height = 22
        r2 += 1

        # Шапка ДЗ
        dz_headers = ["ДЗ", "Наименование", "Норм. база", "Ссылка на пункт", "База для %",
                      "% (ПДЗ)", "Формула расчёта", "Сумма, руб.", "Комментарий"]
        for ci, h in enumerate(dz_headers, 1):
            c = ws2.cell(row=r2, column=ci, value=h)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
            c.alignment = wrap_center; c.border = thin2
        ws2.row_dimensions[r2].height = 28
        r2 += 1

        # Описания ДЗ из НЗ
        dz_comments = {
            "перерывы":  "п.26-27 НЗ: при работе на объектах с режимным доступом (жел. дороги, автодороги, аэродромы, электростанции и пр.) — 25% от СПпз",
            "проезд":    "п.28-36 НЗ: расходы на проезд работников до объекта и обратно. Процент зависит от расстояния (км) и стоимости полевых работ. Таблицы 4-7 НЗ.",
            "организаци": "п.37-39 НЗ, ф.(9): расходы на организацию и ликвидацию полевых работ (мобилизация, аренда, логистика оборудования). Таблица 8 НЗ.",
            "районные":  "п.40 НЗ, ф.(10): доплаты работникам за работу в районах с особыми климатическими условиями (районный коэффициент > 1.0). Приложение к НЗ.",
            "неблагопр": "п.21 НЗ, ф.(4): надбавка за производство работ в неблагоприятный климатический период (зима, дожди). Таблица 3 НЗ.",
        }

        for dz_num, cost in enumerate(estimate.additional_costs, 1):
            ws2.row_dimensions[r2].height = 52
            name_lc = cost.get('name', '').lower()
            comment = next((v for k, v in dz_comments.items() if k in name_lc), "—")

            vals = [
                f"ДЗ-{dz_num}",
                cost.get('name', '—'),
                cost.get('basis', '—'),
                cost.get('basis', '—').split(',')[0] if ',' in cost.get('basis', '') else '—',
                "СПпз (стоимость полевых работ)",
                f"{cost.get('percent', 0):.1f}%",
                cost.get('formula', '—'),
                round(float(cost.get('value', 0)), 2),
                comment,
            ]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=r2, column=ci, value=v)
                c.fill = dz_fill
                c.border = thin2
                c.alignment = wrap_left
                c.font = Font(size=9)
                if ci == 8:  # сумма
                    c.number_format = money_format
                    c.font = Font(bold=True, size=10)
                    c.alignment = wrap_right
            r2 += 1

        # Итого ДЗ
        dz_total = sum(round(float(c.get('value', 0)), 2) for c in estimate.additional_costs)
        ws2.merge_cells(f'A{r2}:G{r2}')
        c = ws2.cell(row=r2, column=1, value="Итого дополнительных затрат:")
        c.font = Font(bold=True, size=10); c.fill = total2_fill
        c.alignment = Alignment(horizontal='right', vertical='center')
        w2_border_row(r2)
        for ci2 in range(1, 12):
            ws2.cell(row=r2, column=ci2).fill = total2_fill
        c8 = ws2.cell(row=r2, column=8, value=dz_total)
        c8.font = Font(bold=True, size=11); c8.number_format = money_format
        c8.alignment = wrap_right; c8.border = thin2; c8.fill = total2_fill
        ws2.row_dimensions[r2].height = 18
        r2 += 2

    # =========================================================
    # Итоговый блок
    # =========================================================
    base_total2  = sum(float(i.total_cost) for i in estimate.items)
    dz_sum2      = sum(round(float(c.get('value', 0)), 2) for c in (estimate.additional_costs or []))
    total_dz2    = round(base_total2 + dz_sum2, 2)
    idx2         = float(estimate.price_index)
    total_idx2   = round(total_dz2 * idx2, 2)
    k_c2         = float(estimate.contract_coefficient)
    final2       = round(total_idx2 * k_c2, 2)

    summary_rows = [
        ("ИТОГО базовые затраты (СП + СЛ + СК, в ценах на 01.01.2024):", base_total2),
        (f"Сумма дополнительных затрат:", dz_sum2),
        (f"ИТОГО с ДЗ (в ценах на 01.01.2024):", total_dz2),
        (f"× Индекс пересчёта ({idx2:.2f}) → приведение к текущим ценам:", total_idx2),
    ]
    if k_c2 != 1.0:
        summary_rows.append((f"× Коэффициент договорной цены ({k_c2:.3f}):", final2))
    summary_rows.append(("ВСЕГО ПО СМЕТЕ (итоговая договорная стоимость):", final2))

    for label, val in summary_rows:
        ws2.merge_cells(f'A{r2}:H{r2}')
        c = ws2.cell(row=r2, column=1, value=label)
        is_total = "ВСЕГО" in label or "ИТОГО с ДЗ" in label
        c.font = Font(bold=True, size=10 if not is_total else 12)
        c.fill = total2_fill if is_total else subtotal_fill
        c.alignment = Alignment(horizontal='right', vertical='center')
        w2_border_row(r2)
        for ci2 in range(1, 12):
            ws2.cell(row=r2, column=ci2).fill = (total2_fill if is_total else subtotal_fill)
        cv = ws2.cell(row=r2, column=9, value=val)
        cv.font = Font(bold=True, size=10 if not is_total else 12)
        cv.number_format = money_format
        cv.alignment = wrap_right
        cv.border = thin2
        cv.fill = total2_fill if is_total else subtotal_fill
        ws2.row_dimensions[r2].height = 20
        r2 += 1

    r2 += 2
    ws2.merge_cells(f'A{r2}:K{r2}')
    ws2.cell(row=r2, column=1,
             value="Смета составлена в соответствии с Приказом Минстроя России от 12.05.2025 № 281/пр. Объёмы работ обоснованы СП 446.1325800.2019, СП 341.1325800.2017, ГОСТ 20522-2012.")
    ws2.cell(row=r2, column=1).font = Font(italic=True, size=9, color="666666")

    # Сохранение
    output_path = Path(filename)
    wb.save(output_path)

    return output_path
